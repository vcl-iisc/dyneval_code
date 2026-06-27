import os
import json
import pandas as pd
import argparse
from collections import defaultdict

def auto_adjust_col_width(df, worksheet, index=True):
    """Automatically adjust Excel column width"""
    from openpyxl.utils import get_column_letter
    idx_cols = df.index.nlevels if index else 0
    for i, col in enumerate(df.columns, idx_cols + 1):
        max_len = max(
            [len(str(col[0])) if isinstance(col, tuple) else len(str(col)),  # First row
             len(str(col[1])) if isinstance(col, tuple) else 0] +           # Second row
            [len(str(cell)) for cell in df.iloc[:, i - idx_cols - 1].values]
        )
        col_letter = get_column_letter(i)
        worksheet.column_dimensions[col_letter].width = max_len + 2
    # Adjust index column width
    if index:
        worksheet.column_dimensions['A'].width = max(df.index.map(lambda x: len(str(x))).max() + 2, 8)

def bold_max_in_each_column(df, worksheet):
    """Bold the max value in each column, including Mean_Accuracy. Auto-adapt for multi-level header and index."""
    from openpyxl.styles import Font

    n_header = df.columns.nlevels
    n_index = df.index.nlevels
    n_rows = df.shape[0]
    n_cols = df.shape[1]
    for col_idx in range(n_cols):
        col = df.iloc[:, col_idx]
        # Only consider non-null and float-convertible values
        col_float = pd.to_numeric(col, errors='coerce')
        max_val = col_float.max()
        if pd.isnull(max_val):
            continue
        for row_idx, val in enumerate(col_float):
            if val == max_val:
                # Row number = header levels + 2 + row index (because pandas writes an extra blank row)
                cell = worksheet.cell(row=n_header + 2 + row_idx, column=n_index + 1 + col_idx)
                cell.font = Font(bold=True)

def main():
    parser = argparse.ArgumentParser(description="Count model accuracy by attribute, output to Excel with multiple sheets")
    parser.add_argument('--input_dir', type=str, required=True, help='Root directory of eval_results')
    args = parser.parse_args()

    EVAL_RESULTS_DIR = args.input_dir
    OUTPUT_EXCEL = os.path.join(EVAL_RESULTS_DIR, 'result_summary.xlsx')

    # First collect all attributes and types
    all_attr_types = set()
    model_data = defaultdict(lambda: defaultdict(dict))  # model -> attribute -> type -> (acc, total, correct)

    print(f"Start traversing directory: {EVAL_RESULTS_DIR}")
    for model in os.listdir(EVAL_RESULTS_DIR):
        model_path = os.path.join(EVAL_RESULTS_DIR, model)
        if not os.path.isdir(model_path):
            continue

        print(f"\nProcessing model: {model}")
        for attribute in os.listdir(model_path):
            attribute_path = os.path.join(model_path, attribute)
            if not os.path.isdir(attribute_path):
                continue

            print(f"  Attribute: {attribute}")
            for length_type in os.listdir(attribute_path):
                length_path = os.path.join(attribute_path, length_type)
                if not os.path.isdir(length_path):
                    continue

                print(f"    Type: {length_type}")
                total_questions = 0
                correct_answers = 0
                file_count = 0

                for filename in os.listdir(length_path):
                    if not filename.endswith('.json'):
                        continue
                    file_path = os.path.join(length_path, filename)
                    file_count += 1
                    try:
                        with open(file_path, 'r', encoding='utf-8') as f:
                            data = json.load(f)
                            gt_answers = data.get('gt_answers', [])
                            model_pred = data.get('model_pred', [])
                            for gt, pred in zip(gt_answers, model_pred):
                                total_questions += 1
                                if gt.strip().lower() == pred.strip().lower():
                                    correct_answers += 1
                    except Exception as e:
                        print(f"      Error reading file: {file_path}, Error: {e}")

                if total_questions > 0:
                    accuracy = correct_answers / total_questions
                else:
                    accuracy = None

                print(f"      File count: {file_count}, Total questions: {total_questions}, Correct answers: {correct_answers}, Accuracy: {accuracy if accuracy is not None else 'N/A'}")

                all_attr_types.add((attribute, length_type))
                model_data[model][attribute][length_type] = {
                    'accuracy': accuracy,
                    'total': total_questions,
                    'correct': correct_answers,
                }

    # Unified column order
    all_attr_types = sorted(all_attr_types)  # [(attribute, type), ...]
    attributes = sorted(set(attr for attr, _ in all_attr_types))
    types = sorted(set(t for _, t in all_attr_types))

    # Build MultiIndex columns
    col_tuples = []
    for attr in attributes:
        for t in types:
            col_tuples.append((attr, t))
    columns = pd.MultiIndex.from_tuples(col_tuples, names=["Attribute", "Type"])

    # Build DataFrame
    def build_df(metric):
        rows = []
        for model in sorted(model_data.keys()):
            row = []
            for attr, t in col_tuples:
                value = model_data[model].get(attr, {}).get(t, {}).get(metric, None)
                row.append(value)
            rows.append(row)
        df = pd.DataFrame(rows, index=sorted(model_data.keys()), columns=columns)
        # Add Mean_Accuracy column (only for accuracy table)
        if metric == 'accuracy':
            df['Mean_Accuracy'] = df.mean(axis=1, skipna=True)
        return df

    accuracy_df = build_df('accuracy')
    count_df = build_df('total')
    correct_df = build_df('correct')

    # Keep two decimal places
    accuracy_df = accuracy_df.applymap(lambda x: f"{x:.4f}" if x is not None and x != "" else "")
    count_df = count_df.applymap(lambda x: f"{x:.4f}" if x is not None and x != "" else "")
    correct_df = correct_df.applymap(lambda x: f"{x:.4f}" if x is not None and x != "" else "")

    # Sort by Mean_Accuracy
    sort_index = pd.to_numeric(accuracy_df['Mean_Accuracy'], errors='coerce').sort_values(ascending=False).index
    accuracy_df = accuracy_df.loc[sort_index]
    count_df = count_df.loc[sort_index]
    correct_df = correct_df.loc[sort_index]

    print("\nWriting to Excel file...")
    with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
        accuracy_df.to_excel(writer, sheet_name='Accuracy')
        count_df.to_excel(writer, sheet_name='Total_Questions')
        correct_df.to_excel(writer, sheet_name='Correct_Answers')

        # Only bold max value in Accuracy sheet
        worksheet = writer.sheets['Accuracy']
        auto_adjust_col_width(accuracy_df, worksheet, index=True)
        bold_max_in_each_column(accuracy_df, worksheet)

        # Only adjust column width for the other two sheets
        for sheet_name, df in zip(['Total_Questions', 'Correct_Answers'], [count_df, correct_df]):
            worksheet = writer.sheets[sheet_name]
            auto_adjust_col_width(df, worksheet, index=True)

    print(f"\nAll statistics have been saved to {OUTPUT_EXCEL}, including three sheets: Accuracy, Total_Questions, and Correct_Answers.")
    print("Done!")

if __name__ == '__main__':
    main()
